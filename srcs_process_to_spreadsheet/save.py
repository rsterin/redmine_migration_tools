from datetime import datetime
from srcs_process_to_spreadsheet import config, logger
import pandas as pd
import openpyxl, os

def export_projects(projects, cleaned_path):
	"""
	Saves multiple projects to individual Excel files.

	Args:
		projects (list): List of project dictionaries.
		cleaned_path (str): Path to save the cleaned Excel files.

	Returns:
		None
	"""
	for project in projects:
		project_name = project["name"].replace(" ", "_")
		project_output_path = os.path.join(cleaned_path, f"{project_name}.xlsx")
		logger.info(f"Preparing to save project {project['name']} to {project_output_path}.")

		try:
			with pd.ExcelWriter(project_output_path, engine="xlsxwriter") as writer:
				export_project_details(writer, project)
				export_issues(writer, project)
				export_memberships(writer, project)
				export_versions(writer, project)
				export_files(writer, project)
				export_time_entries(writer, project)

			apply_excel_formatting(project_output_path)
			logger.info(f"Saved project {project['name']} to {project_output_path}.")
			print(f"Saved project {config.BOLD}{project['name']}{config.END} to {config.BOLD}{project_output_path}{config.END}.")
		except Exception as e:
			logger.error(f"Error while saving project {project['name']} to {project_output_path}: {e}", exc_info=True)
			print(config.BOLD + "Error:\n" + config.END + f"{e}")

def export_project_details(writer, project):
	"""
	Saves project details to an Excel sheet.

	Args:
		writer (pd.ExcelWriter): Pandas ExcelWriter object.
		project (dict): Project dictionary containing project details.

	Returns:
		None
	"""
	project_details = {
		"ID": [project.get("id")],
		"Name": [project.get("name")],
		"Identifier": [project.get("identifier")],
		"Description": [project.get("description")],
		"Status": [project.get("status")],
		"Is Public": [project.get("is_public")],
		"Created On": [format_date(project.get("created_on"))],
		"Updated On": [format_date(project.get("updated_on"))],
	}
	pd.DataFrame(project_details).to_excel(writer, sheet_name="Project Details", index=False, startrow=1)
	worksheet_details = writer.sheets["Project Details"]
	cell_format = writer.book.add_format({
		'text_wrap': True,
		'align': 'center',
		'valign': 'vcenter'
	})
	merge_format = writer.book.add_format({
		'text_wrap': True,
		'align': 'center',
		'valign': 'vcenter',
		'bold': True,
		'border': 1
	})
	worksheet_details.set_column('A:H', None, cell_format)
	for row_num, content in enumerate(pd.DataFrame(project_details).values, start=2):
		worksheet_details.set_row(row_num, None, cell_format)
	worksheet_details.merge_range('A1:H1', 'Project info', merge_format)
	worksheet_details.freeze_panes(2, 0)

def export_issues(writer, project):
	"""
	Saves project issues to an Excel sheet.

	Args:
		writer (pd.ExcelWriter): Pandas ExcelWriter object.
		project (dict): Project dictionary containing project issues.

	Returns:
		None
	"""
	if "issues" in project:
		issues = []
		row_mapping = {}
		relations = []

		for index, issue in enumerate(project["issues"], start=2):
			issue_id = issue.get("id")
			parent_id = issue.get("parent", {}).get("id")
			row_mapping[issue_id] = index

			# Prepare issue data
			issue_data = {
				"ID": issue.get("id"),
				"Project ID": issue.get("project", {}).get("id"),
				"Project Name": issue.get("project", {}).get("name"),
				"Author ID": issue.get("author", {}).get("id"),
				"Author Name": issue.get("author", {}).get("name"),
				"Tracker Name": issue.get("tracker", {}).get("name"),
				"Status Name": issue.get("status", {}).get("name"),
				"Priority Name": issue.get("priority", {}).get("name"),
				"Parent ID": issue.get("parent", {}).get("id"),
				"Subject": issue.get("subject"),
				"Description": issue.get("description"),
				"Start Date": format_date(issue.get("start_date")),
				"Due Date": format_date(issue.get("due_date")),
				"Done Ratio": issue.get("done_ratio"),
				"Estimated Hours": issue.get("estimated_hours"),
				"Created On": format_date(issue.get("created_on")),
				"Updated On": format_date(issue.get("updated_on")),
			}

			relation_columns = []
			if issue and "relations" in issue and issue.get("relations") is not None:
				tmp = issue.get("relations", {}).get("relations", [])
				for relation in tmp:
					relation_columns.append({
						"ID": relation.get("id"),
						"Issue ID": relation.get("issue_id"),
						"Issue To ID": relation.get("issue_to_id"),
						"Relation Type": relation.get("relation_type"),
						"Delay": relation.get("delay"),
						"index": index
					})
			relations.append(relation_columns)
			issues.append(issue_data)

		if issues:
			pd.DataFrame(issues).to_excel(writer, sheet_name="Issues", index=False, startrow=1)
			worksheet_issues = writer.sheets["Issues"]
			parent_col_idx = pd.DataFrame(issues).columns.get_loc("Parent ID")

			for row_num, issue in enumerate(issues, start=2):
				parent_id = issue.get("Parent ID")
				if parent_id and parent_id in row_mapping:
					parent_row = row_mapping[parent_id]
					link = f'internal:Issues!A{parent_row}'
					worksheet_issues.write_url(row_num, parent_col_idx, link, string=str(parent_id))

			cell_format = writer.book.add_format({
				'text_wrap': True,
				'align': 'center',
				'valign': 'vcenter'
			})
			merge_format = writer.book.add_format({
				'text_wrap': True,
				'align': 'center',
				'valign': 'vcenter',
				'bold': True,
				'border': 1
			})
			worksheet_issues.set_column('A:Q', None, cell_format)
			for row_num, content in enumerate(pd.DataFrame(issues).values, start=2):
				worksheet_issues.set_row(row_num, None, cell_format)

			worksheet_issues.merge_range('A1:C1', 'Project info', merge_format)
			worksheet_issues.merge_range('D1:E1', 'Author info', merge_format)
			worksheet_issues.merge_range('F1:Q1', 'Issue info', merge_format)

			max_col = 0
			for relation in relations:
				if relation:
					col_index = pd.DataFrame(issues).columns.get_loc("Updated On") + 1
					for relation_index in relation:
						worksheet_issues.write(1, col_index, "ID", merge_format)
						worksheet_issues.write(1, col_index + 1, "Issue ID", merge_format)
						worksheet_issues.write(1, col_index + 2, "Issue To ID", merge_format)
						worksheet_issues.write(1, col_index + 3, "Relation Type", merge_format)
						worksheet_issues.write(1, col_index + 4, "Delay", merge_format)

						issue_id_row = row_mapping.get(relation_index.get("Issue ID")) + 1
						issue_to_id_row = row_mapping.get(relation_index.get("Issue To ID")) + 1

						worksheet_issues.write(relation_index.get("index"), col_index, relation_index.get("ID"), cell_format)
						if issue_id_row:
							link = f'internal:Issues!A{issue_id_row}'
							worksheet_issues.write_url(relation_index.get("index"), col_index + 1, link, string=str(relation_index.get("Issue ID")))
						else:
							worksheet_issues.write(relation_index.get("index"), col_index + 1, relation_index.get("Issue ID"), cell_format)

						if issue_to_id_row:
							link = f'internal:Issues!A{issue_to_id_row}'
							worksheet_issues.write_url(relation_index.get("index"), col_index + 2, link, string=str(relation_index.get("Issue To ID")))
						else:
							worksheet_issues.write(relation_index.get("index"), col_index + 2, relation_index.get("Issue To ID"), cell_format)

						worksheet_issues.write(relation_index.get("index"), col_index + 3, relation_index.get("Relation Type"), cell_format)
						worksheet_issues.write(relation_index.get("index"), col_index + 4, relation_index.get("Delay"), cell_format)

						if max_col < col_index:
							max_col = col_index
						col_index += 5
			if max_col > 0:
				col_index = pd.DataFrame(issues).columns.get_loc("Updated On") + 1
				index = 1
				while max_col >= col_index:
					worksheet_issues.merge_range(0, col_index, 0, col_index + 4, f'Relation {index}', merge_format)
					col_index += 5
					index += 1

			worksheet_issues.freeze_panes(2, 0)

def export_memberships(writer, project):
	"""
	Saves project memberships to an Excel sheet.

	Args:
		writer (pd.ExcelWriter): Pandas ExcelWriter object.
		project (dict): Project dictionary containing project memberships.

	Returns:
		None
	"""
	if "memberships" in project and "memberships" in project["memberships"]:
		memberships = [
			{
				"Membership ID": m["id"],
				"User ID": m["user"]["id"] if "user" in m else m["anonymous_user"]["id"],
				"User Name": m["user"]["name"] if "user" in m else m["anonymous_user"]["name"],
				"Role": ", ".join([role["name"] for role in m["roles"]]),
			}
			for m in project["memberships"]["memberships"]
		]
		if memberships:
			pd.DataFrame(memberships).to_excel(writer, sheet_name="Memberships", index=False, startrow=1)
			worksheet_memberships = writer.sheets["Memberships"]
			cell_format = writer.book.add_format({
				'text_wrap': True,
				'align': 'center',
				'valign': 'vcenter'
			})
			merge_format = writer.book.add_format({
				'text_wrap': True,
				'align': 'center',
				'valign': 'vcenter',
				'bold': True,
				'border': 1
			})
			worksheet_memberships.set_column('A:D', None, cell_format)
			for row_num, content in enumerate(pd.DataFrame(memberships).values, start=2):
				worksheet_memberships.set_row(row_num, None, cell_format)
			worksheet_memberships.merge_range('A1:D1', 'Memberships info', merge_format)
			worksheet_memberships.freeze_panes(2, 0)

def export_versions(writer, project):
	"""
	Saves project versions to an Excel sheet.

	Args:
		writer (pd.ExcelWriter): Pandas ExcelWriter object.
		project (dict): Project dictionary containing project versions.

	Returns:
		None
	"""
	if "versions" in project and "versions" in project["versions"]:
		versions = [
			{
				"Version ID": v["id"],
				"Name": v["name"],
				"Description": v["description"],
				"Status": v["status"],
				"Due Date": format_date(v["due_date"]),
				"Created On": format_date(v["created_on"]),
				"Updated On": format_date(v["updated_on"]),
			}
			for v in project["versions"]["versions"]
		]
		if versions:
			pd.DataFrame(versions).to_excel(writer, sheet_name="Versions", index=False, startrow=1)
			worksheet_versions = writer.sheets["Versions"]
			cell_format = writer.book.add_format({
				'text_wrap': True,
				'align': 'center',
				'valign': 'vcenter'
			})
			merge_format = writer.book.add_format({
				'text_wrap': True,
				'align': 'center',
				'valign': 'vcenter',
				'bold': True,
				'border': 1
			})
			worksheet_versions.set_column('A:D', None, cell_format)
			for row_num, content in enumerate(pd.DataFrame(versions).values, start=2):
				worksheet_versions.set_row(row_num, None, cell_format)
			worksheet_versions.merge_range('A1:G1', 'Version info', merge_format)
			worksheet_versions.freeze_panes(2, 0)

def export_files(writer, project):
	"""
	Saves project files to an Excel sheet.

	Args:
		writer (pd.ExcelWriter): Pandas ExcelWriter object.
		project (dict): Project dictionary containing project files.

	Returns:
		None
	"""
	if "files" in project and "files" in project["files"]:
		files = [
			{
				"File ID": f["id"],
				"Filename": f["filename"],
				"Filesize": f["filesize"],
				"Content Type": f["content_type"],
				"Description": f["description"],
				"Download URL": f["content_url"],
				"Created On": format_date(f["created_on"]),
				"Downloads": f["downloads"],
				"Author ID": f["author"]["id"],
				"Author Name": f["author"]["name"],
			}
			for f in project["files"]["files"]
		]
		if files:
			pd.DataFrame(files).to_excel(writer, sheet_name="Files", index=False, startrow=1)
			worksheet_files = writer.sheets["Files"]
			cell_format = writer.book.add_format({
				'text_wrap': True,
				'align': 'center',
				'valign': 'vcenter'
			})
			merge_format = writer.book.add_format({
				'text_wrap': True,
				'align': 'center',
				'valign': 'vcenter',
				'bold': True,
				'border': 1
			})
			worksheet_files.set_column('A:J', None, cell_format)
			for row_num, content in enumerate(pd.DataFrame(files).values, start=2):
				worksheet_files.set_row(row_num, None, cell_format)
			worksheet_files.merge_range('A1:H1', 'File info', merge_format)
			worksheet_files.merge_range('I1:J1', 'Author info', merge_format)
			worksheet_files.freeze_panes(2, 0)

def export_time_entries(writer, project):
	"""
	Saves project time entries to an Excel sheet.

	Args:
		writer (pd.ExcelWriter): Pandas ExcelWriter object.
		project (dict): Project dictionary containing project time entries.

	Returns:
		None
	"""
	time_entries = []
	for issue in project.get("issues", []):
		for time_entry in issue.get("time_entries", []):
			time_entry_data = {
				"ID": time_entry.get("id"),
				"Project ID": time_entry.get("project", {}).get("id"),
				"Project Name": time_entry.get("project", {}).get("name"),
				"Issue ID": time_entry.get("issue", {}).get("id"),
				"User ID": time_entry.get("user", {}).get("id"),
				"User Name": time_entry.get("user", {}).get("name"),
				"Activity ID": time_entry.get("activity", {}).get("id"),
				"Activity Name": time_entry.get("activity", {}).get("name"),
				"Hours": time_entry.get("hours"),
				"Comments": time_entry.get("comments"),
				"Spent On": format_date(time_entry.get("spent_on")),
				"Created On": format_date(time_entry.get("created_on")),
				"Updated On": format_date(time_entry.get("updated_on")),
			}
			time_entries.append(time_entry_data)

	if time_entries:
		pd.DataFrame(time_entries).to_excel(writer, sheet_name="Time Entries", index=False, startrow=1)
		worksheet_time_entries = writer.sheets["Time Entries"]
		cell_format = writer.book.add_format({
			'text_wrap': True,
			'align': 'center',
			'valign': 'vcenter'
		})
		merge_format = writer.book.add_format({
			'text_wrap': True,
			'align': 'center',
			'valign': 'vcenter',
			'bold': True,
			'border': 1
		})
		worksheet_time_entries.set_column('A:M', None, cell_format)
		for row_num, content in enumerate(pd.DataFrame(time_entries).values, start=2):
			worksheet_time_entries.set_row(row_num, None, cell_format)
		worksheet_time_entries.merge_range('A1:D1', 'Project info', merge_format)
		worksheet_time_entries.merge_range('E1:F1', 'User info', merge_format)
		worksheet_time_entries.merge_range('G1:M1', 'Time Entry info', merge_format)
		worksheet_time_entries.freeze_panes(2, 0)

def apply_excel_formatting(project_output_path):
	"""
	Formats the Excel file to improve readability.

	Args:
		project_output_path (str): Path to the Excel file to be formatted.

	Returns:
		None
	"""
	wb = openpyxl.load_workbook(f"{project_output_path}")

	cell_format = openpyxl.styles.Alignment(
		horizontal='center',
		vertical='center',
		wrap_text=True
	)

	for sheet_name in wb.sheetnames:
		ws = wb[sheet_name]

		ws.row_dimensions[1].height = 15

		column_widths = {}
		for cell in ws[2]:
			column = cell.column_letter
			max_length = len(str(cell.value))
			adjusted_width = (max_length + 4) * 1.2
			ws.column_dimensions[column].width = adjusted_width
			column_widths[column] = adjusted_width

		for row in ws.iter_rows():
			max_height = 0
			for cell in row:
				cell.alignment = cell_format
				if cell.value:
					char_count = len(str(cell.value))

					column = cell.column_letter
					column_width = column_widths.get(column, 10)

					estimated_height = (char_count // column_width) * 15 + 20
					max_height = max(max_height, estimated_height)
			ws.row_dimensions[cell.row].height = max(15, max_height)

		max_row = ws.max_row
		max_col = ws.max_column
		ws.auto_filter.ref = f"A2:{openpyxl.utils.get_column_letter(max_col)}{max_row}"
	wb.save(f"{project_output_path}")

def format_date(date_str):
	"""
	Formats a date string into a more readable format.

	Args:
		date_str (str): Date string to be formatted.

	Returns:
		str: Formatted date string.
	"""
	try:
		date_obj = datetime.strptime(date_str, "%Y-%m-%dT%H:%M:%SZ")
		return date_obj.strftime("%d-%m-%Y %H:%M:%S")
	except ValueError:
		try:
			date_obj = datetime.strptime(date_str, "%Y-%m-%d")
			return date_obj.strftime("%d-%m-%Y")
		except ValueError:
			return date_str
